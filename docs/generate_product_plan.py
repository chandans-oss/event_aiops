"""
AIOps Platform — Product Planning Excel Generator (Refined v2.0)
Teams: Frontend, Backend, R&D
Timeline: 3 Months (12 Weeks) — Week-by-Week
Includes: Gantt Chart, Timeline, Module Master, Work Items, Benchmarks, Roadmap
Run: python generate_product_plan.py
Output: AIOps_Product_Plan.xlsx
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

# ── Palette ──────────────────────────────────────────────────────────────────
C = {
    "dark_blue": "1E3A5F", "mid_blue": "2563EB", "light_blue": "DBEAFE",
    "dark_green": "065F46", "light_green": "D1FAE5",
    "dark_orange": "92400E", "light_orange": "FEF3C7",
    "dark_red": "7F1D1D", "light_red": "FEE2E2",
    "dark_purple": "4C1D95", "light_purple": "EDE9FE",
    "dark_teal": "134E4A", "light_teal": "CCFBF1",
    "gray_hdr": "374151", "light_gray": "F3F4F6", "white": "FFFFFF",
    "yellow": "FEF08A", "amber": "F59E0B",
    "gantt_fe": "3B82F6", "gantt_be": "10B981", "gantt_rd": "F59E0B",
    "gantt_mile": "EF4444", "gantt_bg": "F8FAFC", "gantt_bar_bg": "E2E8F0",
}

def fill(color): return PatternFill("solid", fgColor=color)
def font(color="111827", bold=False, size=9, italic=False):
    return Font(color=color, bold=bold, size=size, italic=italic)
def align(h="left", v="top", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def border():
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(ws, r, c, val, bg="1E3A5F", fg="FFFFFF", sz=10, bold=True, h="center"):
    cell = ws.cell(row=r, column=c, value=val)
    cell.fill = fill(bg); cell.font = font(fg, bold, sz)
    cell.alignment = align(h, "center"); cell.border = border()
    return cell

def cel(ws, r, c, val, bg=None, fg="111827", bold=False, sz=9, h="left", italic=False):
    cell = ws.cell(row=r, column=c, value=val)
    if bg: cell.fill = fill(bg)
    cell.font = font(fg, bold, sz, italic)
    cell.alignment = align(h, "top"); cell.border = border()
    return cell

def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

def row_height(ws, r, h): ws.row_dimensions[r].height = h

def title_row(ws, title, cols, bg="1E3A5F"):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    c = ws.cell(row=1, column=1, value=title)
    c.fill = fill(bg); c.font = Font(color="FFFFFF", bold=True, size=14)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

# ═══════════════════════════════════════════════════════════════════
# THREE TEAMS DEFINITION
# ═══════════════════════════════════════════════════════════════════
TEAMS = [
    ("Frontend Team", "Angular 14/TypeScript UI Development",
     "Dashboard, Events, Admin Pages, Playground UI, RCA Sidebar, Remediation Sidebar, Auth UI, Topology Viz",
     "3–4 Engineers",
     "Angular 14, TypeScript, RxJS, Chart.js, Angular Material, WebSocket/SSE, CSS/SCSS"),
    ("Backend Team", "Django/FastAPI Core APIs, Infrastructure & Services",
     "Event Ingestion, Kafka, Dedup/Suppression Engines, Correlation Engine, RCA Pipeline APIs, KB CRUD, Remediation Execution, Auth/RBAC, API Gateway, TSDB, Vector DB, CI/CD",
     "4–5 Engineers",
     "Django, Django REST Framework, FastAPI, Kafka, Redis, PostgreSQL, Docker, K8s, Nginx/Kong"),
    ("R&D Team", "AI/ML Research, Model Training, LLM Integration & Innovation",
     "Qwen 3B Serving, RF/IF/KMeans Models, Granger Causality, Sequence Mining, GNN Correlation, Fine-Tuning, Failure Chain Learning, XAI, Drift Detection, Vector Embeddings",
     "3–4 Engineers/Researchers",
     "PyTorch, scikit-learn, ONNX, vLLM, sentence-transformers, PyG/DGL, LoRA/QLoRA, MLflow"),
]

# ═══════════════════════════════════════════════════════════════════
# MODULE MASTER (Team column uses only Frontend / Backend / R&D)
# ═══════════════════════════════════════════════════════════════════
MODULES = [
    ("M01","Dashboard","AIOps Landing Dashboard","Main KPI overview: RCA Insights, SLA Breach Risk, Capacity Risk, Anomalies, Recommended Actions.","P1","✅ Done (Mock)","API Integration Needed","Phase 1","Frontend"),
    ("M02","Dashboard","Topology Dashboard","Node-graph: device dependency chains, alarm propagation, event flow stepper.","P1","✅ Done (Mock)","Real topology API + WebSocket","Phase 2","Frontend + Backend"),
    ("M03","Dashboard","ROI / KPI Dashboard","MTTR reduction, automation rate, cost avoidance, SLA compliance trends.","P2","✅ Done (Static)","Finance API integration","Phase 3","Frontend + Backend"),
    ("M04","Events","Event List & Filtering","Paginated event table with multi-field filters.","P1","✅ Done (Mock)","Real NMS/SNMP event feed","Phase 1","Frontend + Backend"),
    ("M05","Events","Event Actions","Per-event actions: RCA & Impact sidebar, Probable Cause chain.","P1","✅ Done","Link to live RCA API","Phase 1","Frontend"),
    ("M06","Admin","Platform Configuration","Global settings: thresholds, tenant config, integrations.","P1","⚠️ Placeholder","Full config API + RBAC","Phase 2","Backend"),
    ("M07","Admin","Knowledge Base (KB)","KB entries with vector embeddings. View/edit individual records.","P1","✅ Done (Mock)","Vector DB write/read API","Phase 1","Backend + R&D"),
    ("M08","Admin","Correlation Patterns","PatternGallery: view, enable/disable learned correlation rules.","P1","✅ Done (Mock)","Pattern registry API","Phase 2","Backend + R&D"),
    ("M09","Admin","ML Training Config","Scope selection, algorithm selection, duration config.","P1","✅ Done (Mock)","Real training job submission API","Phase 2","Frontend + R&D"),
    ("M10","Admin","Training Analysis","Per-model terminal streaming for all ML algorithms.","P1","✅ Done (Simulated)","Real training job status API","Phase 2","Frontend + R&D"),
    ("M11","Admin","Training Results","List completed training runs; filter by scope/date.","P2","✅ Done (Local)","DB-backed results store","Phase 3","Backend"),
    ("M12","Admin","Training Reports","Full drill-down per run: algorithm outputs, metrics.","P2","✅ Done (Static Data)","Dynamic report from model artifacts","Phase 3","Backend + R&D"),
    ("M13","Playground","RCA Playground (7-Step)","Upload event file → 7-step pipeline → animated results.","P1","✅ Done (API stub)","Full backend pipeline integration","Phase 1","Frontend + Backend + R&D"),
    ("M14","Playground","Event Deduplication Lab","6 dedup techniques interactive demo.","P1","✅ Done (Mock)","Real dedup engine API calls","Phase 2","Frontend + Backend"),
    ("M15","Playground","Event Suppression Lab","12 suppression policies demo.","P1","✅ Done (Mock)","Real suppression engine API","Phase 2","Frontend + Backend"),
    ("M16","Playground","Bulk Event Processing","Batch run suppression + dedup on large event volumes.","P2","⚠️ Partial","File upload API + batch job","Phase 2","Backend"),
    ("M17","Playground","Live Inference Engine","30s poll cycle, 13 ML model artifacts, failure chain matching.","P1","✅ Done (Simulated)","Real model serving + live telemetry","Phase 2","R&D + Backend"),
    ("M18","Core Engine","Event Pre-Processing","Dedup + suppress + normalize raw NMS alarms.","P1","⚠️ Design Only","Production streaming processor (Kafka)","Phase 1","Backend"),
    ("M19","Core Engine","Event Correlation Engine","Temporal, Spatial, Topological, Causal, ML/GNN, LLM Semantic.","P1","⚠️ Partial Design","Graph correlation engine","Phase 1","Backend + R&D"),
    ("M20","Core Engine","RCA Orchestration","Incident + Goal creation from correlated cluster.","P1","✅ Done (API stub)","Real incident management","Phase 1","Backend"),
    ("M21","Core Engine","RCA Intent Router","Score each intent by signal threshold + log keywords.","P1","✅ Done (API stub)","Production intent library, LLM scoring","Phase 1","R&D + Backend"),
    ("M22","Core Engine","RCA Hypothesis Scorer","Score hypotheses: signal + log = total. Select top.","P1","✅ Done (API stub)","Dynamic hypothesis expansion","Phase 1","R&D"),
    ("M23","Core Engine","RCA Situation Builder","Build situation card + vector embedding to DB.","P1","✅ Done (API stub)","Vector DB integration","Phase 1","Backend + R&D"),
    ("M24","Core Engine","RCA Planner LLM","Plan tool calls for diagnostic steps (future).","P2","❌ Not Active","Qwen 3B tool-call R&D","Phase 3","R&D"),
    ("M25","Core Engine","RCA Historical Retriever","Vector similarity search in KB. Top-N retrieval.","P1","✅ Done (API stub)","Production vector DB","Phase 1","Backend + R&D"),
    ("M26","Core Engine","RCA Correlator LLM","Final RCA via Qwen 3B: root cause + confidence + remedy.","P1","✅ Done (API stub)","Qwen 3B production serving","Phase 1","R&D + Backend"),
    ("M27","Core Engine","Remediation Engine","Guided Remediation: Playbook, step states, AI assistant.","P1","✅ Done (Simulated)","Real device execution API","Phase 2","Backend + R&D"),
    ("M28","AI/ML","Random Forest Predictor","150-tree ensemble per device type.","P1","✅ Model Artifacts","ONNX conversion, model serving","Phase 2","R&D"),
    ("M29","AI/ML","Anomaly Detection (IF)","Isolation Forest 5% contamination per device type.","P1","✅ Model Artifacts","Drift detection, auto-retraining","Phase 2","R&D"),
    ("M30","AI/ML","K-Means Clustering","k=4 behavioral clusters per device type.","P1","✅ Model Artifacts","Auto k-selection, drift monitoring","Phase 2","R&D"),
    ("M31","AI/ML","Sequence Mining","Frequent 3-event patterns (support≥2, lift≥1.5).","P1","✅ Done (Static)","Real event stream mining","Phase 2","R&D"),
    ("M32","AI/ML","Cross Correlation","Lagged Pearson & Spearman for metric pairs.","P1","✅ Done (Static)","Real-time sliding window service","Phase 2","R&D"),
    ("M33","AI/ML","Granger Causality","F-test + OLS for directional metric causality.","P1","✅ Done (Static)","Streaming Granger service","Phase 2","R&D"),
    ("M34","AI/ML","Failure Chain Patterns","4 templates for step-by-step chain matching.","P1","✅ Done (Hardcoded)","Dynamic template learning","Phase 2","R&D"),
    ("M35","Infrastructure","SNMP / Syslog Ingestion","Real-time NMS event ingestion pipeline.","P1","❌ Not Built","Kafka + SNMP trap + Syslog parser","Phase 1","Backend"),
    ("M36","Infrastructure","Telemetry Store (TSDB)","Time-series metric storage for ML features.","P1","❌ Not Built","InfluxDB / TimescaleDB","Phase 1","Backend"),
    ("M37","Infrastructure","Vector Database","Situation card embedding storage + KB retrieval.","P1","❌ Not Built","Qdrant / ChromaDB — R&D eval","Phase 1","Backend + R&D"),
    ("M38","Infrastructure","API Gateway","Unified REST API with auth, rate limiting.","P1","❌ Not Built","FastAPI + Kong/Nginx","Phase 1","Backend"),
    ("M39","Infrastructure","Auth & RBAC","Multi-tenant auth, role-based access control.","P1","❌ Not Built","Keycloak / Auth0","Phase 1","Backend"),
    ("M40","Infrastructure","LLM Serving (Qwen 3B)","Local LLM inference for RCA pipeline.","P1","❌ Not Built","vLLM / Ollama, GPU server","Phase 1","R&D + Backend"),
]

# ═══════════════════════════════════════════════════════════════════
# 3-MONTH WEEKLY TIMELINE (12 Weeks)
# ═══════════════════════════════════════════════════════════════════
START_DATE = datetime(2026, 4, 13)  # Monday W1

TIMELINE = [
    # (Week, Phase, Frontend Tasks, Backend Tasks, R&D Tasks, Milestone, Deliverable)
    ("W1", "Phase 1",
     "Auth login page & route guards\nDashboard API integration scaffold",
     "API Gateway setup (FastAPI + Nginx)\nAuth service (JWT + RBAC middleware)\nKafka cluster deployment",
     "Qwen 3B local serving setup (vLLM/Ollama)\nVector DB evaluation PoC (Qdrant vs Chroma)",
     "🏁 Infra Bootstrap", "API Gateway live, Auth working, Kafka running"),

    ("W2", "Phase 1",
     "Event list → real API integration\nRCA Sidebar → wire to cluster API",
     "SNMP trap receiver + Syslog parser v1\nEvent normalization pipeline\nTSDB setup (InfluxDB/TimescaleDB)",
     "Qwen 3B benchmarking (latency, throughput, VRAM)\nVector DB selection finalized",
     "📡 Ingestion Online", "Raw events flowing into Kafka"),

    ("W3", "Phase 1",
     "Dashboard KPI widgets → live API data\nEvent filtering → backend query params",
     "Dedup engine: SHA256 + State Transition\nSuppression engine: Maintenance Window + Flap\nKB CRUD API (GET/POST/PUT/DELETE)",
     "Intent library expansion (10→30 intents)\nHypothesis library expansion\nEmbedding pipeline for KB",
     "🔧 Pre-Processing Engines", "Dedup + Suppression functional"),

    ("W4", "Phase 1",
     "RCA Playground → full backend integration\nKB page → real CRUD forms",
     "Temporal + Causal correlation engine\nRCA Orchestrator API (POST /rca/run)\nRCA Intent Router API",
     "Qwen 3B intent scoring integration\nVector DB write pipeline for situations\nSituation embedding generation",
     "🧠 RCA Pipeline v1", "Steps 0-2 working end-to-end"),

    ("W5", "Phase 1",
     "RCA Playground → animated step results\nEvent actions → live RCA trigger",
     "RCA Hypothesis Scorer API\nRCA Situation Builder API\nVector DB similarity search API",
     "Qwen 3B hypothesis scoring\nHistorical retriever: vector search\nRCA Correlator: final LLM call",
     "🎯 Full RCA Pipeline", "All 7 RCA steps functional"),

    ("W6", "Phase 1",
     "Dashboard real-time refresh (polling)\nError handling & toast notifications\nSidebar nav polish",
     "Parent-Child topology suppression\nRedis caching layer for dedup\nAudit logging for all actions",
     "RCA accuracy validation & tuning\nPrompt engineering for better RCA output\nContext window optimization",
     "✅ Phase 1 Complete", "Core platform demo-ready"),

    ("W7", "Phase 2",
     "Topology Dashboard → dynamic API data\nClickable nodes → event drill-down\nRemediation Sidebar → wire to API",
     "Remediation playbook execution API\nDevice connection pool (SSH/Netconf)\nTopological correlation (graph-based)",
     "Random Forest → ONNX conversion\nIsolation Forest → ONNX conversion\nK-Means → ONNX serving",
     "🔌 Topology + Remediation", "Topology live, Remediation API ready"),

    ("W8", "Phase 2",
     "Training Config → server-side job submit\nTraining Analysis → SSE log streaming\nDedup Lab → real API wiring",
     "Model Training API (POST /train/start)\nTraining Status SSE streaming\nSemantic Dedup (LLM embeddings)",
     "Sequence Mining on real event streams\nCross Correlation sliding window\nGranger Causality streaming service",
     "📊 Training Pipeline", "ML training jobs submittable"),

    ("W9", "Phase 2",
     "Suppression Lab → real engine API\nLive Inference → real model polling\nBulk event processing UI",
     "Suppression: Dynamic Threshold (p95)\nModel Artifact Registry (MLflow)\nRollback engine for remediation",
     "Failure chain dynamic learning\nLive inference: 13 model serving\nConfidence jump logic tuning",
     "🤖 Intelligence Layer", "Live inference operational"),

    ("W10", "Phase 2",
     "Correlation Patterns → enable/disable\nWebSocket/SSE for real-time dashboard\nDark/light mode persistence",
     "Pattern registry API (CRUD + toggle)\nRemedy safety guardrails v1\nEvent bulk-select actions API",
     "GNN correlation cluster refinement\nFeedback loop: operator confirms/rejects RCA\nLLM tool-call evaluation (Planner R&D)",
     "✅ Phase 2 Complete", "Full intelligence + operations layer"),

    ("W11", "Phase 3",
     "ROI/KPI Dashboard → finance API\nTraining Results → DB-backed list\nExport to CSV/PDF",
     "Multi-tenancy data isolation\nRetraining scheduler (drift trigger)\nITSM ticket integration (ServiceNow/Jira)",
     "Fine-tuning Qwen 3B on network domain\nXAI: SHAP values for RF predictions\nModel drift detection pipeline",
     "💼 Business Value", "ROI dashboard, multi-tenancy"),

    ("W12", "Phase 3",
     "Training Reports → dynamic drill-down\nPerformance optimization (Lighthouse)\nFinal UI polish & responsive fixes",
     "Performance tuning (load testing)\nSecurity audit & penetration testing\nDocumentation & API specs",
     "Performance: inference at scale (batching)\nA/B testing framework for models\nFinal model accuracy benchmarking",
     "🚀 Production Ready", "Platform ready for deployment"),
]

# ═══════════════════════════════════════════════════════════════════
# GANTT CHART DATA (task, team, start_week, end_week)
# ═══════════════════════════════════════════════════════════════════
GANTT_TASKS = [
    # --- Frontend ---
    ("Auth Login & Route Guards", "Frontend", 1, 1),
    ("Dashboard API Integration", "Frontend", 1, 3),
    ("Event List & Filtering → API", "Frontend", 2, 3),
    ("RCA Sidebar → Live API", "Frontend", 2, 5),
    ("RCA Playground Integration", "Frontend", 4, 5),
    ("KB CRUD Forms", "Frontend", 4, 4),
    ("Error Handling & Polish", "Frontend", 6, 6),
    ("Topology Dashboard (Dynamic)", "Frontend", 7, 7),
    ("Remediation Sidebar → API", "Frontend", 7, 8),
    ("Training Config & Analysis UI", "Frontend", 8, 9),
    ("Dedup/Suppression Labs → API", "Frontend", 8, 9),
    ("Live Inference UI", "Frontend", 9, 10),
    ("WebSocket/SSE Real-Time", "Frontend", 10, 10),
    ("ROI Dashboard & Reports", "Frontend", 11, 12),
    ("Final Polish & Optimization", "Frontend", 12, 12),
    # --- Backend ---
    ("API Gateway + Auth + Kafka", "Backend", 1, 1),
    ("SNMP/Syslog Ingestion", "Backend", 2, 3),
    ("TSDB Setup & Integration", "Backend", 2, 2),
    ("Dedup Engine (SHA256/State)", "Backend", 3, 3),
    ("Suppression Engine", "Backend", 3, 4),
    ("KB CRUD API", "Backend", 3, 3),
    ("Correlation Engine", "Backend", 4, 5),
    ("RCA Pipeline APIs (Steps 0-6)", "Backend", 4, 5),
    ("Redis Caching + Audit Log", "Backend", 6, 6),
    ("Remediation Execution API", "Backend", 7, 8),
    ("Device Connection Pool", "Backend", 7, 8),
    ("Training API + SSE Streaming", "Backend", 8, 8),
    ("Semantic Dedup + Threshold", "Backend", 8, 9),
    ("Model Registry (MLflow)", "Backend", 9, 9),
    ("Pattern Registry + Safety", "Backend", 10, 10),
    ("Multi-Tenancy + ITSM", "Backend", 11, 11),
    ("Security Audit + Perf Tuning", "Backend", 12, 12),
    # --- R&D ---
    ("Qwen 3B Local Setup", "R&D", 1, 2),
    ("Vector DB Evaluation", "R&D", 1, 2),
    ("Intent & Hypothesis Libraries", "R&D", 3, 4),
    ("Embedding Pipeline", "R&D", 3, 4),
    ("Qwen 3B Scoring Integration", "R&D", 4, 5),
    ("RCA LLM Pipeline Testing", "R&D", 5, 6),
    ("ONNX Model Conversion", "R&D", 7, 7),
    ("Sequence Mining + Cross Corr", "R&D", 8, 8),
    ("Granger Causality Streaming", "R&D", 8, 9),
    ("Failure Chain Learning", "R&D", 9, 9),
    ("Live Inference Serving", "R&D", 9, 10),
    ("GNN Correlation R&D", "R&D", 10, 10),
    ("LLM Fine-Tuning (LoRA)", "R&D", 11, 12),
    ("XAI + Drift Detection", "R&D", 11, 12),
    ("Scale Testing + Benchmarks", "R&D", 12, 12),
]

GANTT_MILESTONES = [
    (1, "🏁 Infra Bootstrap"),
    (2, "📡 Ingestion Online"),
    (4, "🧠 RCA Pipeline v1"),
    (6, "✅ Phase 1 Complete"),
    (7, "🔌 Topology + Remediation"),
    (10, "✅ Phase 2 Complete"),
    (12, "🚀 Production Ready"),
]

# ═══════════════════════════════════════════════════════════════════
# R&D ITEMS
# ═══════════════════════════════════════════════════════════════════
RND_ITEMS = [
    ("R01","LLM — Qwen 3B Serving","P1","HIGH","Deploy Qwen 3B on local GPU (vLLM/Ollama). Benchmark: latency <3s, throughput >5 req/s. Evaluate Q4/Q8.","GPU server (≥24GB VRAM), benchmark suite","W1–W2","R&D"),
    ("R02","LLM — Tool-Call / Function Calling","P2","HIGH","Evaluate Qwen 3B structured tool-call output for Planner LLM step.","Qwen 3B access, incident test dataset","W10","R&D"),
    ("R03","LLM — Domain Fine-Tuning","P2","MEDIUM","Fine-tune on network incident logs/RCA reports. LoRA/QLoRA approach.","GPU cluster, 5000+ labeled pairs","W11–W12","R&D"),
    ("R04","Vector DB Selection","P1","HIGH","Evaluate ChromaDB vs Qdrant vs Weaviate. Criteria: speed, multi-tenancy, scale.","3 PoCs, benchmark tool","W1–W2","R&D + Backend"),
    ("R05","GNN Correlation","P3","HIGH","Graph Neural Network for topology-aware cluster refinement.","GPU, topology dataset, PyG/DGL","W10","R&D"),
    ("R06","Granger at Scale","P2","MEDIUM","Streaming Granger causality: 50+ metrics × 100+ devices in real-time.","Real telemetry, Flink/Spark","W8–W9","R&D"),
    ("R07","Dynamic Failure Chains","P2","HIGH","Auto-discover new failure chain templates from production data.","Event history, mining pipeline","W9","R&D"),
    ("R08","Intent Library Expansion","P1","MEDIUM","Expand from ~10 to 50+ intents covering BGP, OSPF, MPLS, QoS, Security.","Network SMEs, labeled incidents","W3–W4","R&D"),
    ("R09","Remediation Safety","P1","CRITICAL","Dry-run mode, impact assessment, approval workflow, rollback plan.","Security architects","W7–W10","Backend + R&D"),
    ("R10","SNMP/Syslog Normalization","P1","HIGH","Parse 20+ vendor formats to canonical schema.","MIB files, parser framework","W2–W3","Backend"),
    ("R11","Feedback Loop","P2","MEDIUM","Operator confirms/rejects RCA → retrain models. Closed-loop learning.","Feedback UI, retraining pipeline","W10","R&D + Frontend"),
    ("R12","Inference at Scale","P2","HIGH","Batch/cache 13 model calls; ONNX optimization for 100s of devices.","ONNX Runtime, load testing","W12","R&D"),
]

# ═══════════════════════════════════════════════════════════════════
# PERFORMANCE BENCHMARKS
# ═══════════════════════════════════════════════════════════════════
BENCHMARKS = [
    ("Qwen 3B LLM","Single inference latency","N/A (no GPU)","< 3 seconds","< 5s (P95)","curl / wrk2","Q4 quantized recommended"),
    ("Qwen 3B LLM","Throughput (req/s)","N/A","≥ 5 req/s","≥ 3 req/s","locust load test","vLLM batch scheduler"),
    ("Qwen 3B LLM","GPU Memory (VRAM)","N/A","~8 GB (Q4)","< 24 GB","nvidia-smi","RTX 3090/4090 or A100"),
    ("RCA Pipeline (7 Steps)","End-to-end latency","~2s (stub)","< 10 seconds","< 15s (P95)","E2E API timer","Step 6 LLM dominates"),
    ("Event Ingestion","Throughput (events/s)","N/A (mock)","≥ 10,000 evt/s","≥ 5,000 evt/s","Kafka bench","Partitioning critical"),
    ("Dedup Engine","Throughput","N/A","≥ 5,000 evt/s","≥ 2,000 evt/s","JMeter/locust","Redis SET O(1)"),
    ("Correlation Engine","Cluster formation","N/A","< 1 second","< 2 seconds","API timer","15min sliding window"),
    ("Random Forest","Prediction latency","< 50ms","< 10ms (ONNX)","< 20ms","pytest-benchmark","ONNX Runtime"),
    ("Vector DB Search","Top-5 retrieval","N/A","< 50ms","< 100ms","Qdrant bench","1M+ embeddings"),
    ("Live Inference","Poll cycle (30 dev)","~500ms (sim)","< 5 seconds","< 10 seconds","E2E timer","Parallel calls"),
    ("API Gateway","REST latency","N/A","< 200ms","< 500ms (P99)","k6 load test","Excludes LLM"),
    ("Frontend","Page load (LCP)","~1.2s","< 2 seconds","< 3 seconds","Lighthouse","Angular production build"),
]

# ═══════════════════════════════════════════════════════════════════
# ROADMAP
# ═══════════════════════════════════════════════════════════════════
ROADMAP = [
    ("Phase 1","Foundation & Core Pipeline","Weeks 1–6 (6 weeks)",
     "Real event ingestion, dedup, suppression, correlation, full 7-step RCA, KB, vector DB, auth, API gateway",
     "M01,M04,M05,M07,M13,M18–M26,M35–M40","⬜ Not Started"),
    ("Phase 2","Intelligence & Operations","Weeks 7–10 (4 weeks)",
     "Live inference, training pipeline, remediation engine, topology, production labs, model serving",
     "M02,M08–M10,M14–M17,M27–M34","⬜ Not Started"),
    ("Phase 3","Business Value & Hardening","Weeks 11–12 (2 weeks)",
     "ROI dashboard, multi-tenancy, training reports, fine-tuning, security audit, performance tuning",
     "M03,M11,M12,M24,R03,R05","⬜ Not Started"),
]

# ═══════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ═══════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
wb.remove(wb.active)

# ─── Sheet 1: Overview ───────────────────────────────────────────
ws = wb.create_sheet("📋 Overview")
title_row(ws, "AIOps Platform — Product Plan (3 Teams × 12 Weeks)", 6, C["dark_blue"])
ws.row_dimensions[1].height = 40
info = [
    ("Platform","AIOps Event Analytics — Network Operations Intelligence"),
    ("Version","v2.0 Refined (Demo → Production Transition)"),
    ("Date", "April 2026"),
    ("Teams","Frontend (3–4) · Backend (4–5) · R&D (3–4) = 10–13 Engineers"),
    ("Timeline","12 Weeks / 3 Months — April to July 2026"),
    ("Frontend Stack","Angular 14, TypeScript, RxJS, Angular Material, Chart.js"),
    ("Backend Stack","Django, Django REST Framework, FastAPI, Kafka, Redis, PostgreSQL"),
    ("AI/ML Stack","Qwen 3B, scikit-learn, ONNX, PyTorch, vLLM, sentence-transformers"),
    ("Status","Demo complete. Moving to production implementation."),
]
for i, (k, v) in enumerate(info, 3):
    ws.row_dimensions[i+1].height = 22
    cel(ws, i+1, 1, k, bg=C["light_blue"], fg=C["dark_blue"], bold=True, sz=10)
    cel(ws, i+1, 2, v, sz=10)
    for cc in range(3,7): cel(ws, i+1, cc, "")

hdr(ws, 14, 1, "Sheet", bg=C["gray_hdr"]); hdr(ws, 14, 2, "Content", bg=C["gray_hdr"]); hdr(ws, 14, 3, "Purpose", bg=C["gray_hdr"])
sheets_info = [
    ("📋 Overview","This sheet","Platform summary, teams, tech stack"),
    ("🧩 Module Master","All 40 modules","Priority, status, team, phase"),
    ("📅 Timeline","12-week plan","Week-by-week tasks per team"),
    ("📊 Gantt Chart","Visual timeline","Bar chart of all tasks across 12 weeks"),
    ("🔬 R&D Requirements","Research items","Areas needing investigation"),
    ("⚡ Benchmarks","Performance SLAs","Demo vs target vs production"),
    ("👥 Teams","3 team structure","Roles, skills, ownership"),
    ("🗺️ Roadmap","3-phase plan","Phase milestones & deliverables"),
]
for i, (s, c2, p) in enumerate(sheets_info, 15):
    bg = C["light_gray"] if i%2==0 else C["white"]
    cel(ws, i, 1, s, bg=bg, bold=True); cel(ws, i, 2, c2, bg=bg); cel(ws, i, 3, p, bg=bg)
set_widths(ws, {1:30, 2:58, 3:50, 4:20, 5:20, 6:20})

# ─── Sheet 2: Module Master ─────────────────────────────────────
ws2 = wb.create_sheet("🧩 Module Master")
title_row(ws2, "Module Master — All Modules, Priority, Status & Team Ownership", 9, C["dark_blue"])
headers2 = ["ID","Module","Sub-Module","Description","Priority","Demo Status","Prod Status","Phase","Team"]
for c2, h in enumerate(headers2, 1): hdr(ws2, 2, c2, h, bg=C["gray_hdr"])
ws2.freeze_panes = "A3"
row_colors = {
    "Dashboard": C["light_blue"], "Events": C["light_green"],
    "Admin": C["light_purple"], "Playground": C["light_teal"],
    "Core Engine": C["light_orange"], "AI/ML": C["light_red"],
    "Infrastructure": "E5E7EB",
}
for i, row in enumerate(MODULES, 3):
    mod = row[1]; bg = row_colors.get(mod, C["white"])
    status = row[5]
    sfg = C["dark_green"] if "✅" in status else (C["dark_orange"] if "⚠️" in status else C["dark_red"])
    pfg = {"P1": C["dark_red"], "P2": C["dark_orange"], "P3": C["dark_green"]}.get(row[4], "111827")
    for c2, val in enumerate(row, 1):
        if c2 == 4: cel(ws2, i, c2, val, bg=bg, sz=8)
        elif c2 == 5: cel(ws2, i, c2, val, bg=bg, fg=pfg, bold=True, h="center")
        elif c2 == 6: cel(ws2, i, c2, val, bg=bg, fg=sfg, bold=True, sz=8)
        else: cel(ws2, i, c2, val, bg=bg, sz=9)
    ws2.row_dimensions[i].height = 38
set_widths(ws2, {1:6,2:14,3:28,4:50,5:7,6:22,7:28,8:10,9:22})

# ─── Sheet 3: Timeline (12-week) ────────────────────────────────
ws3 = wb.create_sheet("📅 Timeline")
title_row(ws3, "3-Month Timeline — Week-by-Week Plan (Frontend · Backend · R&D)", 7, C["dark_teal"])
h3 = ["Week","Phase","Frontend Tasks","Backend Tasks","R&D Tasks","Milestone","Deliverable"]
for c2, h in enumerate(h3, 1): hdr(ws3, 2, c2, h, bg=C["dark_teal"])
ws3.freeze_panes = "A3"

phase_bg = {"Phase 1": C["light_blue"], "Phase 2": C["light_green"], "Phase 3": C["light_purple"]}
for i, row in enumerate(TIMELINE, 3):
    week_num = int(row[0][1:])
    d = START_DATE + timedelta(weeks=week_num-1)
    date_str = f"{row[0]}  ({d.strftime('%b %d')})"
    bg = phase_bg.get(row[1], C["white"])
    cel(ws3, i, 1, date_str, bg=bg, bold=True, sz=10, h="center")
    cel(ws3, i, 2, row[1], bg=bg, bold=True, sz=9, h="center")
    cel(ws3, i, 3, row[2], bg=bg, sz=9)  # Frontend
    cel(ws3, i, 4, row[3], bg=bg, sz=9)  # Backend
    cel(ws3, i, 5, row[4], bg=bg, sz=9)  # R&D
    cel(ws3, i, 6, row[5], bg=bg, bold=True, sz=9, h="center")
    cel(ws3, i, 7, row[6], bg=bg, sz=9)
    ws3.row_dimensions[i].height = 72
set_widths(ws3, {1:16, 2:10, 3:45, 4:48, 5:45, 6:24, 7:38})

# ─── Sheet 4: Gantt Chart ────────────────────────────────────────
ws4 = wb.create_sheet("📊 Gantt Chart")
gantt_cols = 2 + 12  # Task + Team + 12 weeks
title_row(ws4, "Gantt Chart — 12-Week Project Timeline (Frontend · Backend · R&D)", gantt_cols, C["dark_blue"])

# Header row
hdr(ws4, 2, 1, "Task", bg=C["gray_hdr"])
hdr(ws4, 2, 2, "Team", bg=C["gray_hdr"])
for w in range(1, 13):
    d = START_DATE + timedelta(weeks=w-1)
    hdr(ws4, 2, w+2, f"W{w}\n{d.strftime('%b %d')}", bg=C["gray_hdr"], sz=8)
ws4.freeze_panes = "C3"

team_gantt_colors = {"Frontend": C["gantt_fe"], "Backend": C["gantt_be"], "R&D": C["gantt_rd"]}
team_gantt_light = {"Frontend": C["light_blue"], "Backend": C["light_green"], "R&D": C["light_orange"]}

r = 3
# Add a section header, then tasks for each team
for team_name in ["Frontend", "Backend", "R&D"]:
    # Section header
    ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=gantt_cols)
    hdr(ws4, r, 1, f"── {team_name} Team ──", bg=team_gantt_colors[team_name], sz=11, h="left")
    ws4.row_dimensions[r].height = 24
    r += 1

    team_tasks = [t for t in GANTT_TASKS if t[1] == team_name]
    for task_name, team, sw, ew in team_tasks:
        cel(ws4, r, 1, task_name, bg=C["white"], sz=8, bold=True)
        cel(ws4, r, 2, team, bg=team_gantt_light[team], sz=8, h="center")
        for w in range(1, 13):
            col = w + 2
            if sw <= w <= ew:
                cel(ws4, r, col, "█████", bg=team_gantt_colors[team], fg=team_gantt_colors[team], sz=8, h="center")
            else:
                cel(ws4, r, col, "", bg=C["gantt_bg"], sz=8)
        ws4.row_dimensions[r].height = 20
        r += 1

# Milestone row
r += 1
ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
hdr(ws4, r, 1, "MILESTONES", bg=C["gantt_mile"], sz=10, h="left")
for w in range(1, 13):
    ms = [m[1] for m in GANTT_MILESTONES if m[0] == w]
    val = ms[0] if ms else ""
    bg_c = C["light_red"] if val else C["gantt_bg"]
    cel(ws4, r, w+2, val, bg=bg_c, bold=True, sz=7, h="center")
ws4.row_dimensions[r].height = 28

# Legend
r += 2
cel(ws4, r, 1, "Legend:", bold=True, sz=10)
r += 1
cel(ws4, r, 1, "█████", bg=C["gantt_fe"], fg=C["gantt_fe"], sz=9)
cel(ws4, r, 2, "Frontend", bold=True, sz=9)
r += 1
cel(ws4, r, 1, "█████", bg=C["gantt_be"], fg=C["gantt_be"], sz=9)
cel(ws4, r, 2, "Backend", bold=True, sz=9)
r += 1
cel(ws4, r, 1, "█████", bg=C["gantt_rd"], fg=C["gantt_rd"], sz=9)
cel(ws4, r, 2, "R&D", bold=True, sz=9)
r += 1
cel(ws4, r, 1, "◆", bg=C["light_red"], fg=C["dark_red"], sz=9)
cel(ws4, r, 2, "Milestone", bold=True, sz=9)

set_widths(ws4, {1: 36, 2: 12})
for w in range(1, 13):
    ws4.column_dimensions[get_column_letter(w+2)].width = 14

# ─── Sheet 5: R&D Requirements ──────────────────────────────────
ws5 = wb.create_sheet("🔬 R&D Requirements")
title_row(ws5, "R&D Requirements — Research, PoC & Unknowns", 8, C["dark_orange"])
h5 = ["ID","Area","Priority","Risk","Description","Resources","Timeline","Owner"]
for c2, h in enumerate(h5, 1): hdr(ws5, 2, c2, h, bg=C["dark_orange"])
ws5.freeze_panes = "A3"
risk_colors = {"CRITICAL": C["light_red"], "HIGH": C["light_orange"], "MEDIUM": C["yellow"], "LOW": C["light_green"]}
for i, row in enumerate(RND_ITEMS, 3):
    bg = risk_colors.get(row[3], C["white"])
    for c2, val in enumerate(row, 1):
        cel(ws5, i, c2, val, bg=bg, sz=9)
    ws5.row_dimensions[i].height = 45
set_widths(ws5, {1:6,2:34,3:7,4:10,5:60,6:40,7:10,8:16})

# ─── Sheet 6: Benchmarks ────────────────────────────────────────
ws6 = wb.create_sheet("⚡ Benchmarks")
title_row(ws6, "Performance Benchmarks — Demo vs Target vs SLA", 7, C["dark_teal"])
h6 = ["Component","Metric","Demo Value","Target (Prod)","SLA","Method","Notes"]
for c2, h in enumerate(h6, 1): hdr(ws6, 2, c2, h, bg=C["dark_teal"])
ws6.freeze_panes = "A3"
for i, row in enumerate(BENCHMARKS, 3):
    bg = C["light_teal"] if "Qwen" in row[0] else (C["light_gray"] if i%2==0 else C["white"])
    for c2, val in enumerate(row, 1):
        cel(ws6, i, c2, val, bg=bg, sz=9, bold=("Qwen" in row[0] and c2==1))
    ws6.row_dimensions[i].height = 28
set_widths(ws6, {1:28,2:30,3:20,4:20,5:16,6:24,7:32})

# ─── Sheet 7: Teams ─────────────────────────────────────────────
ws7 = wb.create_sheet("👥 Teams")
title_row(ws7, "Team Structure — 3 Teams: Frontend · Backend · R&D", 5, C["dark_purple"])
h7 = ["Team","Focus Area","Modules Owned","Members","Required Skills"]
for c2, h in enumerate(h7, 1): hdr(ws7, 2, c2, h, bg=C["dark_purple"])
ws7.freeze_panes = "A3"
team_colors = [C["light_blue"], C["light_green"], C["light_orange"]]
for i, row in enumerate(TEAMS, 3):
    bg = team_colors[(i-3) % 3]
    for c2, val in enumerate(row, 1):
        cel(ws7, i, c2, val, bg=bg, sz=9, bold=(c2==1))
    ws7.row_dimensions[i].height = 65
set_widths(ws7, {1:20,2:32,3:60,4:16,5:65})

# ─── Sheet 8: Roadmap ───────────────────────────────────────────
ws8 = wb.create_sheet("🗺️ Roadmap")
title_row(ws8, "Production Roadmap — 3 Phases / 12 Weeks", 6, C["dark_blue"])
h8 = ["Phase","Name","Duration","Key Milestones","Modules","Status"]
for c2, h in enumerate(h8, 1): hdr(ws8, 2, c2, h, bg=C["dark_blue"])
ws8.freeze_panes = "A3"
phase_clr = [C["light_red"], C["light_orange"], C["light_green"]]
for i, row in enumerate(ROADMAP, 3):
    bg = phase_clr[(i-3) % 3]
    for c2, val in enumerate(row, 1):
        cel(ws8, i, c2, val, bg=bg, sz=10, bold=(c2 in [1,2]))
    ws8.row_dimensions[i].height = 70
set_widths(ws8, {1:8,2:28,3:20,4:60,5:40,6:16})

# ─── Save ─────────────────────────────────────────────────────────
output = "AIOps_Product_Plan.xlsx"
wb.save(output)
print(f"✅ Excel generated: {output}")
print(f"   Sheets: {[ws.title for ws in wb.worksheets]}")
print(f"   Teams: Frontend · Backend · R&D")
print(f"   Timeline: 12 weeks (3 months)")
print(f"   Total tasks in Gantt: {len(GANTT_TASKS)}")
